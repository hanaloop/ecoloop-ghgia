import os
from typing_extensions import Buffer
import openpyxl
import pandas as pd

class FileUtils:
    async def read_to_pd(self, file_type: str, file: Buffer = None, path: str = None, sheet = None, starting_row = 0, starting_col = 0, last_row = None, last_col = None) -> pd.DataFrame:
        """
        Reads a file into a pandas dataframe. Generic implementation, finds file type automatically.
        """
        if not file and not path:
            raise Exception("Either file or path must be provided")
        
        file = file or path
        extension = file_type.split(".")[-1]
        
        if extension == "csv":
            return await self.read_csv_to_pd(file)
        elif extension in ["xlsx", "xls"]:
            return await self.read_excel_to_pd(file, sheet, starting_row, starting_col, last_row, last_col)
        elif extension == "xml":
            return await self.read_xml_to_pd(file)
        else:
            raise Exception(f"Unsupported file type: {extension}")


    async def read_excel_to_pd(self, file: Buffer | str, sheet = None, starting_row = 0, starting_col = 0, last_row = None, last_col = None) -> pd.DataFrame: ##TODO: Surely this can be simplified
        if sheet is None:
            starting_row, last_row, cols_to_read, nrows = self.calc_read_range(starting_row, starting_col, last_row, last_col)
            df = pd.read_excel(file, skiprows=starting_row, nrows=nrows, usecols=cols_to_read).fillna("")
            df.replace("",None,inplace=True)
            return df
        starting_row, last_row, cols_to_read, nrows = self.calc_read_range(starting_row, starting_col, last_row, last_col)
        df = pd.read_excel(file, sheet_name=sheet, skiprows=starting_row, nrows=nrows, usecols=cols_to_read).fillna("")
        df.replace("",None,inplace=True)
        return df


    async def read_xml_to_pd(self, file: Buffer | str) -> pd.DataFrame:
        ##does it read csv? TODO:
        df = pd.read_xml(file).fillna("")
        df.replace("",None,inplace=True)
        return df

    async def read_csv_to_pd(self, file: Buffer | str) -> pd.DataFrame:
        df = pd.read_csv(file).fillna("")
        df.replace("",None,inplace=True)
        return df
    
    def get_excel_sheet_names(self, file: Buffer | str, path: str) -> list:
        # Check the file extension
        _, file_extension = os.path.splitext(path)
        print(file_extension, file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm'])
        if file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            # Use openpyxl for .xlsx files
            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
        elif file_extension in ['.xls', '.xlt', '.xla']:
            # Use xlrd for .xls files
            workbook = pd.ExcelFile(file)
            sheet_names = workbook.sheet_names
        else:
            raise ValueError("Unsupported file format")

        return sheet_names


    def get_file_extension(self, data_source: str) -> str:
        """
        Returns the file extension of a given data source.

        Parameters:
            data_source (str): The path or filename of the data source.

        Returns:
            str: The file extension of the data source.
        """
        return os.path.splitext(data_source)[1]

    
    def count_excel_rows(self, path, count_header=True)->int:
        """
        Counts the number of rows in an Excel file.

        Args:
            path (str): The path to the Excel file.
            count_header (bool, optional): Whether to include the header row in the count. Defaults to True.

        Returns:
            int: The number of rows in the Excel file.
        """
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        count = 0
        for row in ws:
            if not all([cell.value is None for cell in row]):
                count += 1
        if not count_header:
            count -= 1
        return count

    def calc_read_range(self, starting_row = 0, starting_col = 0, last_row = None, last_col = None)->tuple:
        if starting_col and last_col and starting_col > last_col:
            raise Exception("Starting column cannot be greater than last column")
        if starting_col and not last_col or not starting_col and last_col:
            raise Exception("Both starting and last column must be provided")
        if not starting_row and last_row:
            raise Exception("Starting row must be provided if last row is provided")
        col_range = range(starting_col, last_col) if last_col and starting_col else None
        nrows = last_row - starting_row + 1 if last_row else None
        return starting_row, last_row, col_range, nrows

        