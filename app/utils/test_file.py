import pytest
import pandas as pd
from io import StringIO
from app.utils.file import FileUtils 
import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from app.utils.file import FileUtils

file = FileUtils()
@pytest.mark.asyncio
async def test_read_csv_to_pd():
    mock_file_content = "col1,col2\nval1,val2"
    mock_file = StringIO(mock_file_content)

    file_utils = FileUtils()
    with pytest.raises(Exception):
        # Test with neither file nor path
        await file_utils.read_to_pd(file_type="csv")

    with pytest.raises(Exception):
        # Test with both file and path
        await file_utils.read_to_pd(file_type="csv", file=mock_file, path="dummy/path")

    # Test with only file
    df = await file_utils.read_csv_to_pd(mock_file)
    assert isinstance(df, pd.DataFrame)
    assert not df.empty

@pytest.mark.asyncio
async def test_read_excel_to_pd(mocker):
    mocker.patch("pandas.read_excel", return_value=pd.DataFrame())
    file_utils = FileUtils()
    df = await file_utils.read_excel_to_pd("dummy/path")
    assert isinstance(df, pd.DataFrame)

@pytest.mark.asyncio
async def test_read_xml_to_pd(mocker):
    mocker.patch("pandas.read_xml", return_value=pd.DataFrame())
    file_utils = FileUtils()
    df = await file_utils.read_xml_to_pd("dummy/path")
    assert isinstance(df, pd.DataFrame)

@pytest.mark.asyncio
async def test_read_to_pd_file_type_routing(mocker):
    mocker.patch("pandas.read_csv", return_value=pd.DataFrame())
    mocker.patch("pandas.read_excel", return_value=pd.DataFrame())
    mocker.patch("pandas.read_xml", return_value=pd.DataFrame())

    mock_file = StringIO("dummy content")

    file_utils = FileUtils()

    # Test CSV
    df_csv = await file_utils.read_to_pd(file_type="csv", file=mock_file)
    assert isinstance(df_csv, pd.DataFrame)

    # Test XLSX
    df_xlsx = await file_utils.read_to_pd(file_type="xlsx", path="dummy/path")
    assert isinstance(df_xlsx, pd.DataFrame)

    # Test XML
    df_xml = await file_utils.read_to_pd(file_type="xml", path="dummy/path")
    assert isinstance(df_xml, pd.DataFrame)

    # Test for invalid file type
    with pytest.raises(Exception):
        await file_utils.read_to_pd(file_type="invalid", path="dummy/path")

def test_empty_file(tmp_path):
    # Test when the excel file is empty
    path = tmp_path / "test.xlsx"
    wb = Workbook()
    wb.save(path)
    assert file.count_excel_rows(str(path)) == 0

def test_single_row(tmp_path):
    # Test when the excel file has a single non-empty row
    path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([1, 2, 3])
    wb.save(path)
    assert file.count_excel_rows(str(path)) == 1

def test_multiple_rows(tmp_path):
    # Test when the excel file has multiple rows, some empty and some non-empty
    path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([1, 2, 3])
    ws.append([None, None, None])
    ws.append([4, 5, 6])
    wb.save(path)
    assert file.count_excel_rows(str(path)) == 2
